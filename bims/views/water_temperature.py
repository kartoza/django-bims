from datetime import datetime

import openpyxl
from bims.models.basemap_layer import BaseMapLayer
import logging
from bims.utils.get_key import get_key

from bims.models.location_site import LocationSite
from django.contrib import messages
from django.http import JsonResponse
from bims.models import WaterTemperature, UploadSession
from django.views.generic import TemplateView
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from bims.views.mixin.session_form.mixin import SessionFormMixin
from django.shortcuts import get_object_or_404
from django.http import HttpResponseRedirect, Http404

logger = logging.getLogger('bims')

LOGGING_INTERVAL = [0.5, 1, 2, 3, 24]


class WaterTemperatureView(TemplateView, SessionFormMixin):
    """View for water temperature form"""
    template_name = 'water_temperature_form.html'
    location_site = None
    category = 'water_temperature'
    upload_task = None

    def get_context_data(self, **kwargs):

        context = super(WaterTemperatureView, self).get_context_data(**kwargs)
        if not self.location_site:
            return context
        context['geoserver_public_location'] = get_key(
            'GEOSERVER_PUBLIC_LOCATION')
        context['location_site_name'] = self.location_site.name
        context['location_site_code'] = self.location_site.site_code
        context['location_site_lat'] = self.location_site.get_centroid().y
        context['location_site_long'] = self.location_site.get_centroid().x
        context['site_id'] = self.location_site.id
        context['logging_interval'] = LOGGING_INTERVAL

        try:
            context['bing_key'] = BaseMapLayer.objects.get(source_type='bing').key
        except BaseMapLayer.DoesNotExist:
            context['bing_key'] = ''

        return context

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        site_id = request.GET.get('siteId', None)
        if site_id:
            self.location_site = get_object_or_404(
                LocationSite,
                pk=site_id
            )
        else:
            raise Http404()

        return super(WaterTemperatureView, self).get(request, *args, **kwargs)

    def extra_post(self, post):
        """
        Override this method to process the POST request.
        :param post: POST request
        """
        return

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        owner_id = request.POST.get('owner_id', '').strip()
        interval = request.POST.get('interval')
        xlsx_file = request.FILES.get('xlsx_file')

        upload_session = UploadSession.objects.create(
            uploader=request.user,
            process_file=xlsx_file,
            uploaded_at=datetime.now(),
            category=self.category
        )
        if self.upload_task:
            self.upload_task.delay(upload_session.id)

        work_book = openpyxl.load_workbook(upload_session.process_file, data_only=True)
        work_sheet = work_book.active
        row_number = work_sheet.max_row
        column_number = work_sheet.max_column
        date_start = work_sheet.cell(3, 1)
        date_end = work_sheet.cell(row_number, 1)

        days = date_end.value - date_start.value
        rows = int(days * float(interval))
        if interval == 24:
            if column_number < 4:
                return JsonResponse({
                    'status': 'failed',
                    'message': 'Missing minimum amd maximum data value'
                })

        if rows > row_number:
            return JsonResponse({
                'status': 'failed',
                'message': 'Missing {} rows'.format(rows - row_number - 1)
            })
        elif rows < row_number:
            return JsonResponse({
                'status': 'failed',
                'message': 'Got more {} rows'.format(rows - row_number - 1)
            })
        else:
            for i in row_number - 1:
                if column_number > 2:
                    mean = work_sheet.cell(i, 2).value
                    minimum = work_sheet.cell(i, 3).value
                    maximum = work_sheet.cell(i, 4).value
                    value = work_sheet.cell(i, 2).value
                else:
                    mean = minimum = maximum = value = None
                WaterTemperature.objects.create(
                    date_time=work_sheet.cell(i, 1).value,
                    location_site=request.POST.get('site-id', None),
                    is_daily=False,
                    mean=mean,
                    minimum=minimum,
                    maximum=maximum,
                    value=value,
                    owner=owner_id,
                    source_file=upload_session.process_file
                )
            return JsonResponse({
                'status': 'success',
                'message': 'Water temperature date being uploaded. Thank you'
            })
